import csv
import io
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, Http404
from django.db.models import Count, Avg, Sum, Q
from django.utils.dateparse import parse_datetime
from django.utils import timezone

# Excel and PDF exports
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from .models import User, Course, Student, ActivityLog, Prediction
from .forms import UserRegistrationForm, UserEditForm, CSVUploadForm
from .machine_learning import run_ml_pipeline


# Custom Decorators for Role-Based Access Control
def admin_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if not request.user.is_admin():
            messages.error(request, "Access denied. Admin privileges required.")
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


def instructor_or_admin_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if not (request.user.is_admin() or request.user.is_instructor()):
            messages.error(request, "Access denied. Instructor or Admin privileges required.")
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


def advisor_or_admin_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if not (request.user.is_admin() or request.user.is_advisor()):
            messages.error(request, "Access denied. Advisor or Admin privileges required.")
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


# 1. Authentication Views
def register_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, f"Account created for {user.username}! You can now login.")
            return redirect('login')
        else:
            messages.error(request, "Registration failed. Please check the form errors.")
    else:
        form = UserRegistrationForm()
    return render(request, 'analytics/register.html', {'form': form})


def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f"Welcome back, {user.username} ({user.get_role_display()})!")
                return redirect('dashboard')
            else:
                messages.error(request, "Invalid username or password.")
        else:
            messages.error(request, "Invalid username or password.")
    else:
        form = AuthenticationForm()
    return render(request, 'analytics/login.html', {'form': form})


def logout_view(request):
    logout(request)
    messages.info(request, "You have been logged out.")
    return redirect('login')


# 2. General Dashboard View (Accessible to all authenticated roles)
@login_required
def dashboard_view(request):
    total_students = Student.objects.count()
    total_courses = Course.objects.count()
    total_logs = ActivityLog.objects.count()
    total_predictions = Prediction.objects.count()

    # Risk distribution chart data
    risk_distribution = (
        Prediction.objects.values('predicted_risk')
        .annotate(count=Count('id'))
        .order_by('predicted_risk')
    )
    
    risk_data = {
        'low': 0,
        'medium': 0,
        'high': 0
    }
    for item in risk_distribution:
        r_type = item['predicted_risk']
        if r_type == 'Low Risk':
            risk_data['low'] = item['count']
        elif r_type == 'Medium Risk':
            risk_data['medium'] = item['count']
        elif r_type == 'High Risk':
            risk_data['high'] = item['count']

    # Course Performance data (Top 5 courses by average engagement score)
    course_performance = (
        Prediction.objects.values('course__course_code')
        .annotate(avg_score=Avg('engagement_score'))
        .order_by('-avg_score')[:5]
    )
    course_labels = [item['course__course_code'] for item in course_performance]
    course_scores = [round(item['avg_score'], 1) for item in course_performance]

    # Activity count trends (Daily action counts for the last 10 days with activity)
    # Using SQLite date extractor
    daily_trends = (
        ActivityLog.objects.extra(select={'day': "strftime('%Y-%m-%d', timestamp)"})
        .values('day')
        .annotate(count=Count('id'))
        .order_by('-day')[:10]
    )
    daily_trends = list(daily_trends)[::-1] # Reverse to chrono order
    trend_labels = [item['day'] for item in daily_trends]
    trend_counts = [item['count'] for item in daily_trends]

    # Recent activity log table
    recent_logs = ActivityLog.objects.select_related('student', 'course').order_by('-timestamp')[:8]

    # At-risk students summary list (High Risk)
    at_risk_predictions = Prediction.objects.select_related('student', 'course').filter(
        predicted_risk='High Risk'
    ).order_by('engagement_score')[:6]

    context = {
        'total_students': total_students,
        'total_courses': total_courses,
        'total_logs': total_logs,
        'total_predictions': total_predictions,
        'risk_data': risk_data,
        'course_labels': course_labels,
        'course_scores': course_scores,
        'trend_labels': trend_labels,
        'trend_counts': trend_counts,
        'recent_logs': recent_logs,
        'at_risk_predictions': at_risk_predictions,
    }
    return render(request, 'analytics/dashboard.html', context)


# 3. Admin View: Upload LMS Logs
@admin_required
def upload_csv_view(request):
    if request.method == 'POST':
        form = CSVUploadForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = request.FILES['csv_file']
            try:
                data_set = csv_file.read().decode('UTF-8')
                io_string = io.StringIO(data_set)
                reader = csv.DictReader(io_string)

                logs_created = 0
                for row in reader:
                    s_id = row.get('student_id')
                    s_name = row.get('student_name')
                    s_email = row.get('student_email')
                    c_code = row.get('course_code')
                    c_name = row.get('course_name')
                    timestamp_str = row.get('timestamp')
                    action = row.get('action')
                    duration_str = row.get('duration', '0')

                    if not (s_id and s_name and s_email and c_code and c_name and timestamp_str and action):
                        continue

                    course, _ = Course.objects.get_or_create(
                        course_code=c_code,
                        defaults={'course_name': c_name}
                    )

                    student, _ = Student.objects.get_or_create(
                        student_id=s_id,
                        defaults={'name': s_name, 'email': s_email}
                    )

                    # DateTime parsing
                    parsed_dt = parse_datetime(timestamp_str)
                    if not parsed_dt:
                        try:
                            parsed_dt = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S')
                        except ValueError:
                            try:
                                parsed_dt = datetime.strptime(timestamp_str, '%Y-%m-%d')
                            except ValueError:
                                parsed_dt = timezone.now()

                    if parsed_dt and timezone.is_naive(parsed_dt):
                        parsed_dt = timezone.make_aware(parsed_dt)

                    try:
                        duration = int(duration_str)
                    except ValueError:
                        duration = 0

                    ActivityLog.objects.create(
                        student=student,
                        course=course,
                        timestamp=parsed_dt,
                        action=action,
                        duration=duration
                    )
                    logs_created += 1

                # Train ML & generate predictions
                predictions_updated = run_ml_pipeline()

                messages.success(
                    request, 
                    f"Successfully imported {logs_created} logs and updated {predictions_updated} predictions in the Database."
                )
                return redirect('dashboard')

            except Exception as e:
                messages.error(request, f"Failed to process CSV file. Error: {str(e)}")
    else:
        form = CSVUploadForm()
    
    return render(request, 'analytics/upload.html', {'form': form})


# 4. Admin View: User Management
@admin_required
def manage_users_view(request):
    users = User.objects.all().order_by('-date_joined')
    return render(request, 'analytics/user_management.html', {'users': users})


@admin_required
def edit_user_view(request, user_id):
    user_to_edit = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        form = UserEditForm(request.POST, instance=user_to_edit)
        if form.is_valid():
            form.save()
            messages.success(request, f"User '{user_to_edit.username}' has been updated.")
            return redirect('manage_users')
    else:
        form = UserEditForm(instance=user_to_edit)
    return render(request, 'analytics/edit_user.html', {'form': form, 'user_to_edit': user_to_edit})


# 5. Instructor/Advisor: Course Analytics List & Details
@login_required
def courses_view(request):
    # Requires Instructor, Advisor or Admin
    if not (request.user.is_admin() or request.user.is_instructor() or request.user.is_advisor()):
        raise Http404("Access Denied.")

    courses = Course.objects.annotate(
        student_count=Count('activity_logs__student', distinct=True),
        log_count=Count('activity_logs'),
        avg_score=Avg('predictions__engagement_score')
    ).order_by('course_code')

    return render(request, 'analytics/course_analytics.html', {'courses': courses})


@login_required
def course_detail_view(request, course_id):
    if not (request.user.is_admin() or request.user.is_instructor() or request.user.is_advisor()):
        raise Http404("Access Denied.")

    course = get_object_or_404(Course, id=course_id)
    
    # Aggregated stats for the course
    predictions = Prediction.objects.select_related('student').filter(course=course).order_by('-engagement_score')
    
    # Calculate statistics
    avg_score = predictions.aggregate(Avg('engagement_score'))['engagement_score__avg'] or 0
    at_risk_count = predictions.filter(predicted_risk='High Risk').count()
    medium_risk_count = predictions.filter(predicted_risk='Medium Risk').count()
    low_risk_count = predictions.filter(predicted_risk='Low Risk').count()

    # Activity breakdown (actions distribution)
    action_stats = (
        ActivityLog.objects.filter(course=course)
        .values('action')
        .annotate(count=Count('id'))
        .order_by('-count')
    )

    context = {
        'course': course,
        'predictions': predictions,
        'avg_score': round(avg_score, 1),
        'at_risk_count': at_risk_count,
        'medium_risk_count': medium_risk_count,
        'low_risk_count': low_risk_count,
        'action_stats': action_stats,
    }
    return render(request, 'analytics/course_detail.html', context)


# 6. Instructor/Advisor: Student list & search
@login_required
def students_view(request):
    if not (request.user.is_admin() or request.user.is_instructor() or request.user.is_advisor()):
        raise Http404("Access Denied.")

    query = request.GET.get('search', '')
    if query:
        students = Student.objects.filter(
            Q(name__icontains=query) | Q(student_id__icontains=query)
        )
    else:
        students = Student.objects.all()

    # Pre-fetch predictions for efficiency
    students = students.annotate(
        avg_score=Avg('predictions__engagement_score'),
        courses_count=Count('predictions__course', distinct=True)
    ).order_by('name')

    return render(request, 'analytics/students_list.html', {'students': students, 'query': query})


@login_required
def student_detail_view(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    predictions = Prediction.objects.select_related('course').filter(student=student)
    logs = ActivityLog.objects.select_related('course').filter(student=student).order_by('-timestamp')[:50]

    # Calculations
    overall_avg_score = predictions.aggregate(Avg('engagement_score'))['engagement_score__avg'] or 0
    total_duration_spent = logs.aggregate(Sum('duration'))['duration__sum'] or 0
    total_duration_hours = round(total_duration_spent / 3600.0, 1)

    context = {
        'student': student,
        'predictions': predictions,
        'logs': logs,
        'overall_avg_score': round(overall_avg_score, 1),
        'total_duration_hours': total_duration_hours,
        'total_activities': logs.count(),
    }
    return render(request, 'analytics/student_detail.html', context)


# 7. Advisor: Risk Alerts Page
@advisor_or_admin_required
def risk_alerts_view(request):
    # Fetch High and Medium risk predictions
    query = Prediction.objects.select_related('student', 'course').filter(
        predicted_risk__in=['High Risk', 'Medium Risk']
    )

    # 1. Get filter parameters from GET request
    search_query = request.GET.get('search', '').strip()
    course_id = request.GET.get('course', '')
    risk_level = request.GET.get('risk_level', '')
    max_engagement = request.GET.get('max_engagement', '')
    sort_by = request.GET.get('sort_by', '')

    # 2. Apply filters
    if search_query:
        query = query.filter(
            Q(student__name__icontains=search_query) | 
            Q(student__student_id__icontains=search_query)
        )
    if course_id:
        query = query.filter(course_id=course_id)
    if risk_level:
        query = query.filter(predicted_risk=risk_level)
    if max_engagement:
        try:
            query = query.filter(engagement_score__lte=float(max_engagement))
        except ValueError:
            pass

    # 3. Apply sorting
    if sort_by == 'engagement_asc':
        query = query.order_by('engagement_score')
    elif sort_by == 'engagement_desc':
        query = query.order_by('-engagement_score')
    elif sort_by == 'student_name':
        query = query.order_by('student__name')
    else:
        # Default order: Critical first (High Risk, then lowest engagement score)
        query = query.order_by('-predicted_risk', 'engagement_score')

    # Total counts for statistics show aggregate pending items (unfiltered)
    total_alerts = Prediction.objects.filter(predicted_risk__in=['High Risk', 'Medium Risk'])
    high_risk_count = total_alerts.filter(predicted_risk='High Risk').count()
    medium_risk_count = total_alerts.filter(predicted_risk='Medium Risk').count()

    # Dynamic filtered alert count
    filtered_count = query.count()

    # Get list of courses active in predictions to populate course dropdown filter
    active_course_ids = Prediction.objects.filter(
        predicted_risk__in=['High Risk', 'Medium Risk']
    ).values_list('course_id', flat=True).distinct()
    courses = Course.objects.filter(id__in=active_course_ids).order_by('course_code')

    context = {
        'alerts': query,
        'courses': courses,
        'high_risk_count': high_risk_count,
        'medium_risk_count': medium_risk_count,
        'filtered_count': filtered_count,
        # Pass filters back to the template
        'search_query': search_query,
        'selected_course': course_id,
        'selected_risk': risk_level,
        'selected_max_engagement': max_engagement,
        'selected_sort': sort_by,
    }
    return render(request, 'analytics/advisor_alerts.html', context)




# 8. Report Generator Views (Accessible to Advisors, Instructors, and Admins)
@login_required
def export_pdf_report(request):
    if not (request.user.is_admin() or request.user.is_instructor() or request.user.is_advisor()):
        raise Http404("Access Denied.")

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="lms_student_risk_report.pdf"'

    # Document Template setup
    doc = SimpleDocTemplate(
        response,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )

    styles = getSampleStyleSheet()
    
    # Custom styles matching our premium color system
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=colors.HexColor('#004643'),
        spaceAfter=15
    )
    subtitle_style = ParagraphStyle(
        'DocSub',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor('#555555'),
        spaceAfter=25
    )
    cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontSize=9,
        leading=11
    )
    header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontSize=9,
        fontName='Helvetica-Bold',
        textColor=colors.white
    )

    story = []

    # Title & Metadata
    story.append(Paragraph("LMS Activity Logs: Student Risk Report", title_style))
    story.append(Paragraph(
        f"Generated by: {request.user.username} ({request.user.get_role_display()}) | Date: {timezone.now().strftime('%Y-%m-%d %H:%M:%S UTC')}",
        subtitle_style
    ))
    story.append(Spacer(1, 10))

    # Summary Statistics
    stats_data = [
        [
            Paragraph("<b>Total Students:</b>", cell_style), Paragraph(str(Student.objects.count()), cell_style),
            Paragraph("<b>Total Courses:</b>", cell_style), Paragraph(str(Course.objects.count()), cell_style)
        ],
        [
            Paragraph("<b>High Risk Students:</b>", cell_style), Paragraph(str(Prediction.objects.filter(predicted_risk='High Risk').count()), cell_style),
            Paragraph("<b>Medium Risk Students:</b>", cell_style), Paragraph(str(Prediction.objects.filter(predicted_risk='Medium Risk').count()), cell_style)
        ]
    ]
    summary_table = Table(stats_data, colWidths=[120, 100, 120, 100])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F0EDE5')),
        ('PADDING', (0,0), (-1,-1), 8),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LINEBELOW', (0,0), (-1,-1), 1, colors.HexColor('#D3CFC9')),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 20))

    # Main Predictions Table
    predictions = Prediction.objects.select_related('student', 'course').order_by('-predicted_risk', 'engagement_score')
    
    table_data = [[
        Paragraph("Student ID", header_style),
        Paragraph("Student Name", header_style),
        Paragraph("Course Code", header_style),
        Paragraph("Engagement", header_style),
        Paragraph("Risk Level", header_style),
        Paragraph("Intervention Recommendation", header_style)
    ]]

    for p in predictions:
        # Style risk badge colors inside PDF
        if p.predicted_risk == 'High Risk':
            risk_color = colors.HexColor('#8A1F1F')
        elif p.predicted_risk == 'Medium Risk':
            risk_color = colors.HexColor('#9E7611')
        else:
            risk_color = colors.HexColor('#1E5C3B')

        risk_p_style = ParagraphStyle(
            'RiskStyle',
            parent=cell_style,
            fontName='Helvetica-Bold',
            textColor=risk_color
        )

        table_data.append([
            Paragraph(p.student.student_id, cell_style),
            Paragraph(p.student.name, cell_style),
            Paragraph(p.course.course_code, cell_style),
            Paragraph(f"{p.engagement_score:.1f}%", cell_style),
            Paragraph(p.predicted_risk, risk_p_style),
            Paragraph(p.recommendation or "No recommendation.", cell_style)
        ])

    predictions_table = Table(table_data, colWidths=[65, 95, 65, 65, 65, 200])
    predictions_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#004643')),
        ('PADDING', (0,0), (-1,-1), 6),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D3CFC9')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#FAF8F5')])
    ]))

    story.append(predictions_table)
    
    # Build Document
    doc.build(story)
    return response


@login_required
def export_excel_report(request):
    if not (request.user.is_admin() or request.user.is_instructor() or request.user.is_advisor()):
        raise Http404("Access Denied.")

    # Create openpyxl workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Risk Predictions"

    # Set up columns headers
    headers = [
        "Student ID", "Student Name", "Student Email", 
        "Course Code", "Course Name", "Engagement Score (%)", 
        "Predicted Risk", "Recommendation", "Last Updated"
    ]
    ws.append(headers)

    # Style header row
    header_fill = openpyxl.styles.PatternFill(start_color="004643", end_color="004643", fill_type="solid")
    header_font = openpyxl.styles.Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font

    # Add predictions
    predictions = Prediction.objects.select_related('student', 'course').order_by('-predicted_risk', 'engagement_score')
    
    for p in predictions:
        ws.append([
            p.student.student_id,
            p.student.name,
            p.student.email,
            p.course.course_code,
            p.course.course_name,
            p.engagement_score,
            p.predicted_risk,
            p.recommendation or "",
            p.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        ])

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    # Output to HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="lms_student_risk_predictions.xlsx"'
    wb.save(response)
    
    return response
